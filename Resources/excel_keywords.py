import openpyxl
import re

global device_info
device_info = []


def write_full_table_to_excel(path, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Device Info"
    for i, line in enumerate(data, start=1):
        label, value = line.split("\t", 1)
        ws.cell(row=i, column=1, value=label.strip())
        ws.cell(row=i, column=2, value=value.strip())
    wb.save(path)

def write_full_table_to_csv(path):
    with open(path,"w") as f :
        for i, line in enumerate(device_info, start=1):
            label, value = line[0],line[1]
            f.write(f"{label},{value}\n")

def data_reformating_device_info(text):
    if "MAC" in text:
        device_info.append(["MAC",re.findall(r'MAC\s*:\s*([:0-9A-Z]*)',text)[0]])
        return ["MAC",re.findall(r'MAC\s*:\s*([:0-9A-Z]*)',text)[0]]
    if ":" in text:
        data_list= text.split(":")
        if len(data_list) >= 2:
            device_info.append([text.split(":")[0],text.split(":")[1]])
            return [text.split(":")[0],text.split(":")[1]]
    return[""]